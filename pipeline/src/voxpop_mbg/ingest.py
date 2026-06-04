"""Load raw comment data from a directory or single file."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path

from .config import SUPPORTED_EXTENSIONS
from .utils import iter_supported_files, log

Record = dict[str, object]


class IngestError(RuntimeError):
    pass


@dataclass
class IngestResult:
    records: list[Record]
    source_file: Path
    raw_row_count: int
    structural_duplicates: int


def _resolve_file(input_path: Path, file_override: str | None) -> Path:
    if file_override:
        candidate = Path(file_override)
        if not candidate.is_absolute():
            candidate = (input_path if input_path.is_dir() else input_path.parent) / candidate
        if not candidate.exists():
            raise IngestError(f"Specified file does not exist: {candidate}")
        return candidate

    if input_path.is_file():
        return input_path

    if not input_path.exists():
        raise IngestError(
            f"Input path does not exist: {input_path}. "
            "Download the dataset and place the file in data/raw/."
        )

    files = iter_supported_files(input_path)
    if not files:
        raise IngestError(
            f"No supported files found in {input_path}. "
            f"Supported extensions: {', '.join(SUPPORTED_EXTENSIONS)}."
        )
    if len(files) > 1:
        names = "\n  ".join(path.name for path in files)
        raise IngestError(
            "Multiple supported files found. Re-run with --file to choose one:\n  "
            + names
        )
    return files[0]


def _load_json(path: Path) -> list[Record]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if isinstance(data, dict):
        for key in ("data", "comments", "records", "items"):
            if isinstance(data.get(key), list):
                return data[key]
        return [data]
    if isinstance(data, list):
        return [row for row in data if isinstance(row, dict)]
    raise IngestError(f"Unsupported JSON structure in {path.name}")


def _load_jsonl(path: Path) -> list[Record]:
    rows: list[Record] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            if isinstance(obj, dict):
                rows.append(obj)
    return rows


def _load_csv(path: Path) -> list[Record]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _load_xlsx(path: Path) -> list[Record]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise IngestError(
            "Reading .xlsx files requires openpyxl. Install it with "
            "`pip install openpyxl`, or convert the file to CSV/JSON."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return []
    records: list[Record] = []
    for row in rows:
        record = {header[i]: row[i] for i in range(min(len(header), len(row)))}
        records.append(record)
    return records


def _dedupe_by_id(records: list[Record]) -> tuple[list[Record], int]:
    """Collapse rows that repeat the same identifier (a serialization artifact)."""
    id_column = None
    for candidate in ("comment_id", "id", "commentId"):
        if records and candidate in records[0]:
            id_column = candidate
            break
    if id_column is None:
        return records, 0

    seen: set[str] = set()
    unique: list[Record] = []
    duplicates = 0
    for record in records:
        identifier = record.get(id_column)
        key = str(identifier) if identifier not in (None, "") else None
        if key is not None and key in seen:
            duplicates += 1
            continue
        if key is not None:
            seen.add(key)
        unique.append(record)
    return unique, duplicates


def load_records(input_path: Path, file_override: str | None = None) -> IngestResult:
    source = _resolve_file(input_path, file_override)
    suffix = source.suffix.lower()
    log(f"Loading {source.name} ({suffix})")

    if suffix == ".json":
        records = _load_json(source)
    elif suffix == ".jsonl":
        records = _load_jsonl(source)
    elif suffix == ".csv":
        records = _load_csv(source)
    elif suffix == ".xlsx":
        records = _load_xlsx(source)
    else:
        raise IngestError(f"Unsupported file extension: {suffix}")

    raw_count = len(records)
    deduped, structural = _dedupe_by_id(records)
    if structural:
        log(f"Collapsed {structural} repeated identifier rows")
    return IngestResult(
        records=deduped,
        source_file=source,
        raw_row_count=len(deduped),
        structural_duplicates=structural,
    )
